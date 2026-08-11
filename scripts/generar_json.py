#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
generar_json.py
================
Lee la matriz de indicadores por factor (Excel) y produce el JSON que
consume el tablero: data/datos_indicadores.json

Uso:
    python scripts/generar_json.py

Requisitos:
    pip install openpyxl

Estructura esperada de la hoja "Matriz Indicadores":
    N° Factor | Factor | Indicadores | 2020 | 2021 | 2022 | 2023 | 2024 | 2025

Salida (JSON):
    {
      "years": [2020, ..., 2025],
      "factors": [
        {"n": 1, "factor": "...", "indicators": [
            {"name": "...", "values": [v2020, ..., v2025], "pct": true|false}
        ]}
      ]
    }

Valores de texto ("No Aplica", "ND", "No Disponible", "NaN", vacío) se
convierten a null: la serie salta ese año en el gráfico.
"""

import json
import math
import os
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("Falta openpyxl. Instala con: pip install openpyxl")

# --- Rutas (relativas a la raíz del proyecto) ---
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL = os.path.join(BASE, "data", "Matriz_indicadores_por_factor.xlsx")
SALIDA = os.path.join(BASE, "data", "datos_indicadores.json")
TIPOS_JSON = os.path.join(BASE, "data", "tipos_grafico.json")
HOJA = "Matriz Indicadores"
ANIOS = [2020, 2021, 2022, 2023, 2024, 2025]

# Palabras clave que marcan un indicador como porcentaje (se muestra con %)
CLAVES_PCT = [
    "porcentaje", "tasa", "nivel de satisfacción", "aporte relativo",
    "participación", "absorción", "selectividad", "empleabilidad",
    "deserción",
]

# Frases que contienen una palabra porcentual, pero representan magnitudes.
EXCLUSIONES_PCT = [
    "número promedio de semestres",
]


def limpiar(v):
    """Devuelve número (float/int) o None para cualquier texto/vacío."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        if isinstance(v, float) and math.isnan(v):
            return None
        return round(float(v), 6)
    return None  # "No Aplica", "ND", etc.


def es_porcentaje(nombre):
    n = (nombre or "").lower()
    return not any(frase in n for frase in EXCLUSIONES_PCT) and any(
        clave in n for clave in CLAVES_PCT
    )


def cargar_mapa_tipos():
    """
    Carga los metadatos de tipo de gráfico ('chart') y serie dual ('dual')
    desde tipos_grafico.json y/o datos_indicadores.json existente para
    garantizar que nunca se pierdan al regenerar desde el Excel.
    """
    mapa = {}

    # 1. Leer tipos_grafico.json (fuente principal de metadatos de gráficos)
    if os.path.exists(TIPOS_JSON):
        try:
            with open(TIPOS_JSON, "r", encoding="utf-8") as f:
                d = json.load(f)
                for item in d.get("indicadores", []):
                    nombre = item.get("name", "").strip()
                    if nombre:
                        mapa[nombre] = {
                            "chart": item.get("chart", "linea"),
                            "dual": item.get("dual", False),
                        }
        except Exception as e:
            print("Aviso: no se pudo leer tipos_grafico.json: %s" % e)

    # 2. Leer datos_indicadores.json existente (como respaldo secundario)
    if os.path.exists(SALIDA):
        try:
            with open(SALIDA, "r", encoding="utf-8") as f:
                d = json.load(f)
                for factor in d.get("factors", []):
                    for ind in factor.get("indicators", []):
                        nombre = ind.get("name", "").strip()
                        if nombre and nombre not in mapa:
                            mapa[nombre] = {
                                "chart": ind.get("chart", "linea"),
                                "dual": ind.get("dual", False),
                            }
        except Exception:
            pass

    return mapa


def main():
    if not os.path.exists(EXCEL):
        sys.exit("No se encontró el Excel en: " + EXCEL)

    wb = openpyxl.load_workbook(EXCEL, data_only=True)
    if HOJA not in wb.sheetnames:
        sys.exit('No existe la hoja "%s". Hojas: %s' % (HOJA, wb.sheetnames))

    mapa_tipos = cargar_mapa_tipos()

    ws = wb[HOJA]
    filas = list(ws.iter_rows(values_only=True))[1:]  # saltar cabecera

    factores = {}
    for r in filas:
        if r[0] is None:
            continue
        n_factor = int(r[0])
        nombre_factor = r[1]
        nombre_ind = (r[2] or "").strip()
        valores = [limpiar(r[3 + i]) for i in range(len(ANIOS))]

        # Determinar tipo de gráfico y flag dual
        if nombre_ind in mapa_tipos:
            tipo_chart = mapa_tipos[nombre_ind]["chart"]
            es_dual = mapa_tipos[nombre_ind]["dual"]
        elif "(nacional)" in nombre_ind.lower():
            tipo_chart = "linea"
            es_dual = True
        else:
            tipo_chart = "linea"
            es_dual = False

        factores.setdefault(n_factor, {
            "n": n_factor,
            "factor": nombre_factor,
            "indicators": [],
        })
        factores[n_factor]["indicators"].append({
            "name": nombre_ind,
            "values": valores,
            "pct": es_porcentaje(nombre_ind),
            "chart": tipo_chart,
            "dual": es_dual,
        })

    salida = {
        "years": ANIOS,
        "factors": [factores[k] for k in sorted(factores)],
    }

    with open(SALIDA, "w", encoding="utf-8") as f:
        json.dump(salida, f, ensure_ascii=False, indent=None)

    total_ind = sum(len(x["indicators"]) for x in salida["factors"])
    cant_barras = sum(1 for f in salida["factors"] for i in f["indicators"] if i.get("chart") == "barras")
    cant_lineas = sum(1 for f in salida["factors"] for i in f["indicators"] if i.get("chart") == "linea")
    cant_duales = sum(1 for f in salida["factors"] for i in f["indicators"] if i.get("dual"))

    print("OK -> %s" % SALIDA)
    print("Factores: %d | Indicadores: %d (Barras: %d | Línea: %d | Duales: %d)" % (
        len(salida["factors"]), total_ind, cant_barras, cant_lineas, cant_duales
    ))


if __name__ == "__main__":
    main()
